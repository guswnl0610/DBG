# -*- coding: utf-8 -*-
import sys, json
import bluetooth
import os
import time
import openpyxl
from openpyxl import Workbook

#both side
bd_addr240l = "00:18:e4:35:65:95" #240왼쪽
bd_addr240r = "98:D3:A2:FD:78:70" #240오른쪽
bd_addr230l = "00:18:E4:35:14:14" #230왼쪽
bd_addr230r = "00:18:E4:35:17:81" #230오른쪽
bd_addr250l = "00:18:E4:35:1C:52" #250왼쪽
bd_addr250r = "00:18:E4:35:0C:E8" #250오른쪽

bd_addrnew240l = "98:D3:51:FD:91:C3" #240왼쪽
bd_addrnew240r = "98:D3:71:FD:57:66" #250오른쪽

bd_addrnew250l = "98:D3:61:FD:70:C5" #250왼쪽
bd_addrnew250r = "98:D3:51:FD:92:68" #250오른쪽
#btddr250r = "00:18:91:D7:99:69" #이친구는 예비용 아직사용x


#Read data from stdin
def read_in():
    lines = sys.stdin.readlines()
    # Since our input would only be having one line, parse our JSON data from that
    return json.loads(lines[0])
# def jh(name,size):
#     return str(name)+str(size)

def pickSize(size):
    shoes = []
    if(int(size) == 230):
        shoes.append(bd_addr230l)
        shoes.append(bd_addr230r)
        # shoes.append(bd_addr230l)
        # shoes.append(bd_addr230r)
    elif(int(size) == 240):
        shoes.append(bd_addr240l)
        shoes.append(bd_addr240r)
        # shoes.append(bd_addr240r)
        # shoes.append(bd_addr240l)
    else:
        shoes.append(bd_addr250l)
        shoes.append(bd_addr250r)
    return shoes

    
def main():
    #get our data as an array from read_in()
    lines = read_in()
    port = 1
    lshoesize = pickSize(lines[1])[0]
    rshoesize = pickSize(lines[1])[1]

    #왼발 오른발 커넥트 속도가 달라서 길이가 달라지는 것을 방지하기 위함
    leftcolno = 1
    rightcolno = 6
    lengthdiffer = 0

    filename = "ddobak.xlsx"

    if os.path.isfile(filename):
        wb = openpyxl.load_workbook(filename)
    else:
        wb = Workbook()
        ws = wb.active

    sheet1 = wb.create_sheet(title = str(lines[0]))

    #connect bluetooth 1
    sock=bluetooth.BluetoothSocket(bluetooth.RFCOMM)
    sock.connect((lshoesize, port))
    print('1success')
    sock.setblocking(False)
    sock.settimeout(20.0)

    #recieve bluetooth 1
    temp = sock.recv(256)
    print('recv1')

    #connect bluetooth 2
    sock2=bluetooth.BluetoothSocket(bluetooth.RFCOMM)
    sock2.connect((rshoesize, port))
    print('2success')
    sock2.setblocking(False)
    sock2.settimeout(20.0)

    #recieve bluetooth 2
    temp2 = sock2.recv(256)
    print('recv2')

    # print'걸어라!!!!!!!!!!!!!!'
    #get values 1 on list1
    start_time = time.time()
    fsrList = []
    fsrstring = ""

    #get values 2 on list2
    start_time2 = time.time()
    fsrList2 = []
    fsrstring2 = ""

    while(True):
        temp = sock.recv(1024)
        fsrstring += temp
        if(time.time() - start_time > 5):
            break

        temp2 = sock2.recv(1024)
        fsrstring2 += temp2
        if(time.time() - start_time2 > 5):
            break
    sock.close()
    sock2.close()

    end_time = time.time()
    print(end_time - start_time)
    fsrList = fsrstring.split()
    # print(fsrList)


    end_time2 = time.time()
    print(end_time2 - start_time2)
    fsrList2 = fsrstring2.split()
    # print(fsrList2)

    print('Stop walking!')

    lengthdiffer = ( len(fsrList) - len(fsrList2) ) #길이맞춤
    fsrList = fsrList[lengthdiffer:]                #길이맞춤

    #left foot : divide values by time and add to list
    l_fsr1 = []
    l_fsr2 = []
    l_fsr3 = []
    l_fsr4 = []
    l_mpu = [] #yaw

    firstvaluempuleft = 0
    firstvaluempuright = 0

    for j in fsrList:
        i = float(j)
        if (0 <= i and i < 1000):
            l_fsr1.append(i)
        if (1000 <= i and i < 2000):
            l_fsr2.append(i - 1000)
        if (2000 <= i and i < 3000):
            l_fsr3.append(i - 2000)
        if (3000 <= i and i < 4000):
            l_fsr4.append(i - 3000)
        if (6500 <= i and i < 7500):
            l_mpu.append(i - 7000)

    #right foot : divide values by time and add to list

    r_fsr1 = []
    r_fsr2 = []
    r_fsr3 = []
    r_fsr4 = []
    r_mpu = [] #yaw

    for j in fsrList2:
        i = float(j)
        if (0 <= i and i < 1000):
            r_fsr1.append(i)
        if (1000 <= i and i < 2000):
            r_fsr2.append(i - 1000)
        if (2000 <= i and i < 3000):
            r_fsr3.append(i - 2000)
        if (3000 <= i and i < 4000):
            r_fsr4.append(i - 3000)
        if (6500 <= i and i < 7500):
            r_mpu.append(i - 7000)


    firstvaluempuleft = l_mpu[0]
    firstvaluempuright = r_mpu[0]

    for col_index in range(1, 5):
        sheet1.cell(1, column=col_index).value = 'Lfsr'+ str(col_index)
    sheet1.cell(1, column=5).value = 'Lmpu'
    for col_index in range(6, 10):
        sheet1.cell(1, column=col_index).value = 'Rfsr'+ str(col_index-4)
    sheet1.cell(1, column=10).value = 'Rmpu'

    #upload to file by list_L
    for row_index in range(2, 37):
        sheet1.cell(row=row_index, column=1).value = l_fsr1[(row_index)] if l_fsr1[row_index] > 100 else 0

    for row_index in range(2, 37):
        sheet1.cell(row=row_index, column=2).value = l_fsr2[(row_index)] if l_fsr2[row_index] > 100 else 0

    for row_index in range(2, 37):
        sheet1.cell(row=row_index, column=3).value = l_fsr3[(row_index)] if l_fsr3[row_index] > 100 else 0

    for row_index in range(2, 37):
        sheet1.cell(row=row_index, column=4).value = l_fsr4[(row_index)] if l_fsr4[row_index] > 100 else 0

    for row_index in range(2, 37):
        if( (l_mpu[(row_index)] - firstvaluempuleft) < -180 ):
            sheet1.cell(row=row_index, column=5).value = 0 if abs(l_mpu[row_index] - firstvaluempuleft + 360) > 90 else l_mpu[row_index] - firstvaluempuleft + 360

        else:
            sheet1.cell(row=row_index, column=5).value = 0 if abs(l_mpu[row_index] - firstvaluempuleft) > 90 else l_mpu[row_index] - firstvaluempuleft


    #upload to file by list_R
    for row_index in range(2, 37):
        sheet1.cell(row=row_index, column=6).value = r_fsr1[(row_index)] if r_fsr1[row_index] > 100 else 0

    for row_index in range(2, 37):
        sheet1.cell(row=row_index, column=7).value = r_fsr2[(row_index)] if r_fsr2[row_index] > 100 else 0

    for row_index in range(2, 37):
        sheet1.cell(row=row_index, column=8).value = r_fsr3[(row_index)] if r_fsr3[row_index] > 100 else 0

    for row_index in range(2, 37):
        sheet1.cell(row=row_index, column=9).value = r_fsr4[(row_index)] if r_fsr4[row_index] > 100 else 0

    for row_index in range(2, 37):
        if( (r_mpu[(row_index)] - firstvaluempuright) < -180 ):
            sheet1.cell(row = row_index, column=10).value = 0 if abs(r_mpu[row_index] - firstvaluempuright + 360) > 90 else r_mpu[row_index] - firstvaluempuright + 360

        else:
            sheet1.cell(row=row_index, column=10).value = 0 if abs(r_mpu[row_index] - firstvaluempuright) > 90 else r_mpu[(row_index)] - firstvaluempuright


    #두번째 엑셀에 들어갈 값들 계산



    #save file
    wb.save(filename = filename)


    # Sum  of all the items in the providen array
    # rr = []
    # for item in lines:
    #     # total_sum_inArray += item
    #     rr.append(item)
        

    #return the sum to the output stream
    # print jh(lines[0],lines[1])

# Start process
if __name__ == '__main__':
    main()