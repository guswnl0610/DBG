#-*- coding: utf-8 -*-

import sys, json,io
import pandas as pd
from openpyxl import load_workbook
import numpy as np
from sklearn.model_selection import train_test_split
import os
import struct
from sklearn.linear_model import Perceptron
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import accuracy_score
from scipy.fftpack import fft, ifft


filename = "mns.xlsx"

filename2 = "ddobak.xlsx"#trial
filename3 = "mbti.xlsx"

sheets = []
sheets2 = []
sheets3  = []

# 엑셀파일 불러오기
wb = load_workbook(filename)
wb2 = load_workbook(filename2)
wb3 = load_workbook(filename3)


def readCell(sheet):
    temp = []
    for row in sheet:
        for cell in row:
            temp.append(cell.value)
    return temp


def makeFFT(fsrList):
    maxval = fsrList[fsrList.index(max(fsrList))]
    newFsrlist = []
    for i in fsrList:
        try:
            value = float(i) / maxval
        except ZeroDivisionError as e:
            print(e)
            value = 0
        # newFsrlist.append(float(i) / maxval)
        newFsrlist.append(value)

    fourier = fft(newFsrlist)

    freal = []
    for i in fourier:
        freal.append(i.real)

    return freal

#SI용 testfinal

def testfinalRow_s(index):
    workSheet2 = wb2.get_sheet_by_name(sheets2[index])

    # Left
    lfsr1 = workSheet2['A2':'A36']
    lfsr2 = workSheet2['B2':'B36']
    lfsr3 = workSheet2['C2':'C36']
    lfsr4 = workSheet2['D2':'D36']

    # Right
    rfsr1 = workSheet2['F2':'F36']
    rfsr2 = workSheet2['G2':'G36']
    rfsr3 = workSheet2['H2':'H36']
    rfsr4 = workSheet2['I2':'I36']

    # Left1
    list1 = readCell(lfsr1)
    list2 = readCell(lfsr2)
    list3 = readCell(lfsr3)
    list4 = readCell(lfsr4)

    # Right1
    list5 = readCell(rfsr1)
    list6 = readCell(rfsr2)
    list7 = readCell(rfsr3)
    list8 = readCell(rfsr4)
    
   
    left = []
    left.extend(list1)
    left.extend(list2)
    left.extend(list3)
    left.extend(list4)

    right= []
    right.extend(list5)
    right.extend(list6)
    right.extend(list7)
    right.extend(list8)

    alll = []
    alll.extend(left)
    alll.extend(right)
    
    return alll


def testfinalRow_m(index):
    workSheet2 = wb2.get_sheet_by_name(sheets2[index])

    # Left
    lfsr1 = workSheet2['A2':'A36']
    lfsr2 = workSheet2['B2':'B36']
    lfsr3 = workSheet2['C2':'C36']
    lfsr4 = workSheet2['D2':'D36']

    # Right
    rfsr1 = workSheet2['F2':'F36']
    rfsr2 = workSheet2['G2':'G36']
    rfsr3 = workSheet2['H2':'H36']
    rfsr4 = workSheet2['I2':'I36']

    # Left1
    list1 = readCell(lfsr1)
    list2 = readCell(lfsr2)
    list3 = readCell(lfsr3)
    list4 = readCell(lfsr4)

    # Right1
    list5 = readCell(rfsr1)
    list6 = readCell(rfsr2)
    list7 = readCell(rfsr3)
    list8 = readCell(rfsr4)
    
    # Left
    freq1 = makeFFT(list1)
    freq2 = makeFFT(list2)
    freq3 = makeFFT(list3)
    freq4 = makeFFT(list4)

    # Right
    freq5 = makeFFT(list5)
    freq6 = makeFFT(list6)
    freq7 = makeFFT(list7)
    freq8 = makeFFT(list8)

    leftFreq = []
    leftFreq.extend(freq1)
    leftFreq.extend(freq2)
    leftFreq.extend(freq3)
    leftFreq.extend(freq4)

    rightFreq = []
    rightFreq.extend(freq5)
    rightFreq.extend(freq6)
    rightFreq.extend(freq7)
    rightFreq.extend(freq8)

    allFreq = []
    allFreq.extend(leftFreq)
    allFreq.extend(rightFreq)
   
    return allFreq

#양 발 사이 각도를 return해주는 함수
def classifyINorOUT(filename, sheetnamelist):
    inoroutdf = pd.read_excel(filename, sheet_name=sheetnamelist[len(sheetnamelist)-1])
    differ = (inoroutdf.Rmpu - inoroutdf.Lmpu).sum() / 35
    return differ

def main():
    
    #트레이닝 데이터1 : SI만 있는 파일 -> 데이터 프레임으로 real_df
    df = pd.read_excel(filename, index_col=0)
    real_df = df

    #트레이닝 데이터 2 : MBTI만 있는 파일 -> 데이터 프레임으로 real_df2
    df2 = pd.read_excel(filename3, index_col=0)
    real_df2 = df2

    #test(지금 받은 데이터)를 시트에서 가져와서 새로운 데이터 프레임으로 test_real_df
    #test데이터는 한 사람에 대한 것 <- 마지막 시트에서 가져옴
    #있는 시트 다 가져와서 시트이름으로 한 줄씩 붙여서 데이터 프레임 만들기
    for i in wb2.get_sheet_names():
        sheets2.append(i)

    sheetNames2 = wb2.get_sheet_names()
    last_index = len(sheetNames2) - 1

    #이거는 SI용 테스트 데이터로 탈바꿈 
    data = {}
    temp = {sheetNames2[last_index] : testfinalRow_s(last_index)}
    data.update(temp)

    df_test_s = pd.DataFrame(data)
    test_real_df_s = df_test_s.T

    #이거는 MBTI용 테스트 데이터로 탈바꿈 
    data2 = {}
    temp2 = {sheetNames2[last_index] : testfinalRow_m(last_index)}
    data2.update(temp2)

    df_test_m = pd.DataFrame(data2)
    test_real_df_m = df_test_m.T


    #SI 1,2,3,에 대해서 perceptron 돌리기 
    #결과에 따라 comment 생성하기 
    
    comment = u""
    label = ['SI_1','SI_2','SI_3','SI_4']
    fsrlabel = ['fsr1', 'fsr2', 'fsr3', 'fsr4']
    for i in range(0,4):
        labelIndex = label[i]
        x_train , y_train = np.array(real_df.iloc[0:303,0:280], dtype='f8'), np.array(real_df[labelIndex], dtype='f8')
        x_test = np.array(test_real_df_s.iloc[:,:], dtype='f8')

        #에러 때문에..
        

        sc = StandardScaler() #normalization
        sc.fit(x_train)
        x_train_normal = sc.transform(x_train)
        x_test_normal = sc.transform(x_test)

        perceptron = Perceptron(eta0=0.1, max_iter=1000, random_state=1, tol=0.19) #eta0 is learning rate

        perceptron.fit(x_train_normal,y_train)
        p_predict = perceptron.predict(x_test_normal)
       
        print("p_predict : ",p_predict[0])
        if (p_predict[0] ==0):
            comment += fsrlabel[i] +u" 구역은 정상입니다\n"
        elif (p_predict[0] ==1):
            comment += fsrlabel[i] +u" 구역은 왼쪽으로 무게중심을 더 실어서 걷고 계십니다\n"
        else:
            comment += fsrlabel[i] + u" 구역은 오른쪽으로 무게중심을 더 실어서 걷고 계십니다\n"
        comment+=u"*"
    
    #회전
    degree = classifyINorOUT(filename2,sheetNames2)
    comment += u"당신의 발의 벌어진 각도는 "
    if(degree >= 20):
        comment += str(degree)+ u"도로 팔자걸음이십니다.\n"
    elif(degree <= -20):
        comment += str(degree)+ u"도로 안짱걸음이십니다.\n"
    else:
        comment += str(degree)+ u"도로 정상걸음이십니다.\n"
    comment+=u"*"

    #MBTI에 대해서 perceptron 돌리기
    x_train , y_train = np.array(real_df2.iloc[0:303,0:280], dtype='f8'), np.array(real_df2['MBTI'], dtype='f8')
    x_test = np.array(test_real_df_m.iloc[:,:], dtype='f8')

    sc = StandardScaler() #normalization
    sc.fit(x_train)
    x_train_normal = sc.transform(x_train)
    x_test_normal = sc.transform(x_test)

    perceptron = Perceptron(eta0=0.1, max_iter=1000, random_state=1, tol=0.19) #eta0 is learning rate

    perceptron.fit(x_train_normal,y_train)
    p_predict = perceptron.predict(x_test_normal)
    
    if (p_predict[0] ==0):
        # comment += u" 걸음걸이로 본 당신의 성격은 I,T\n"
        comment += u" 걸음걸이로 본 당신의 성격은 혼자 하는 활동을 선호하며 내면에 담고 글쓰는 것을 선호하는 내향형, 그리고 일과 목표, 효율성을 중시하는 객관적이고 논리적인 성향인 사고형입니다."            
    elif (p_predict[0] ==1):
        # comment += u" 걸음걸이로 본 당신의 성격은 I,F\n"
        comment += u" 걸음걸이로 본 당신의 성격은 혼자 하는 활동을 선호하며 내면에 담고 글쓰는 것을 선호하는 내향형, 그리고 대인관계와 사람을 중시하는 공감적인 성향의 감정형입니다."
    elif (p_predict[0] ==2):
        # comment += u" 걸음걸이로 본 당신의 성격은 E,T\n"
        comment += u" 걸음걸이로 본 당신의 성격은 단체 활동을 선호하며 생각을 표출하며 말하기를 선호하는 외향형, 그리고 일과 목표, 효율성을 중시하는 객관적이고 논리적인 성향인 사고형입니다." 
    else:
        # comment += u" 걸음걸이로 본 당신의 성격은 E,F\n"
         comment += u" 걸음걸이로 본 당신의 성격은 단체 활동을 선호하며 생각을 표출하며 말하기를 선호하는 외향형, 그리고 대인관계와 사람을 중시하는 공감적인 성향의 감정형입니다."
    comment+=u"*"
    comment+=str(p_predict[0])
    
    if(degree >= 20):
        comment += u"*o"
    elif(degree <= -20):
        comment += u"*i"
    else:
        comment += u"*n"
    comment+=u"*"


    #SI에서 생성된 comment -> txt 파일로 저장하기    
    with io.open('./public/texts/result.txt', 'w+', encoding='utf-8') as f:
        f.write(comment)
        f.close()



main()